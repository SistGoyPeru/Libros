#!/usr/bin/env python3
"""
Genera archivos .xlsx de ejemplo en codigos/ para excel_contadores,
uno por capítulo, con contenido contable/financiero real.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, IconSetRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import random

BASE = Path(r"C:\Users\alexg\OneDrive\Alex_2026\libros epub")
DEST = BASE / "excel_contadores" / "codigos"

# ── helpers ──
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="0A7E5C")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ACCENT_FILL = PatternFill("solid", fgColor="E8F5E9")
GOLD_FILL = PatternFill("solid", fgColor="FFF8E1")


def style_header(ws, row=1, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)


def add_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 1 — Configuración / Conciliación Bancaria
# ═══════════════════════════════════════════════════════════════════
def cap1(wb):
    ws = wb.active
    ws.title = "Conciliacion Bancaria"
    ws["A1"] = "Fecha"; ws["B1"] = "Referencia"; ws["C1"] = "Descripcion"
    ws["D1"] = "Debe S/"; ws["E1"] = "Haber S/"; ws["F1"] = "Banco S/"; ws["G1"] = "Diferencia"
    data = [
        ["02/01/2025", "CH-001", "Pago proveedor Maderas Perez", 0, 15000, 15000, None],
        ["05/01/2025", "DP-001", "Deposito cliente Import Sur", 35000, 0, 35000, None],
        ["08/01/2025", "CH-002", "Pago alquiler local", 0, 5000, 5000, None],
        ["10/01/2025", "TF-001", "Transferencia proveedor", 0, 8500, 8500, None],
        ["12/01/2025", "DP-002", "Deposito ventas varias", 12000, 0, 12000, None],
        ["15/01/2025", "CH-003", "Pago servicios publicos", 0, 1200, 1200, None],
        ["18/01/2025", "DP-003", "Cobranza factura F001", 22000, 0, 22000, None],
        ["20/01/2025", "CH-004", "Pago nomina empleados", 0, 18000, 18000, None],
        ["22/01/2025", "TF-002", "Transferencia a proveedor China", 0, 32000, 32000, None],
        ["25/01/2025", "DP-004", "Deposito contado", 8500, 0, 8500, None],
        ["28/01/2025", "CH-005", "Pago impuesto SUNAT", 0, 4200, 4200, None],
        ["30/01/2025", "DP-005", "Deposito cliente Fact F002", 16000, 0, 16000, None],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=7).value = f"=D{i}-E{i}"
        ws.cell(row=i, column=7).number_format = '#,##0.00'
    style_header(ws, 1, 7)
    add_border_range(ws, 1, len(data) + 1, 1, 7)
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=6).number_format = '#,##0.00'
    ws["A14"] = "SALDO FINAL LIBROS"; ws["B14"] = "=SUM(D2:D13)-SUM(E2:E13)"
    ws["B14"].number_format = '#,##0.00'
    ws["A15"] = "SALDO FINAL BANCO"; ws["B15"] = "=SUM(F2:F13)"
    ws["B15"].number_format = '#,##0.00'
    ws["A16"] = "DIFERENCIA"; ws["B16"] = "=B14-B15"
    ws["B16"].number_format = '#,##0.00'
    ws["A16"].font = Font(bold=True, color="E74C3C")
    # formato condicional - diferencias
    ws.conditional_formatting.add("G2:G13",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))
    auto_width(ws)


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 2 — Catálogo de Cuentas
# ═══════════════════════════════════════════════════════════════════
def cap2(wb):
    ws = wb.active
    ws.title = "Catalogo Cuentas"
    ws["A1"] = "Codigo"; ws["B1"] = "Descripcion"; ws["C1"] = "Nivel"
    ws["D1"] = "Tipo"; ws["E1"] = "Grupo"; ws["F1"] = "Estado"
    data = [
        ["10.01.01.01", "Efectivo en Caja", 4, "Activo", "Balance", "Activo"],
        ["10.01.01.02", "Caja General Soles", 5, "Activo", "Balance", "Activo"],
        ["10.01.02.01", "Banco SCotiabank Cta Cte", 4, "Activo", "Balance", "Activo"],
        ["10.01.02.02", "Banco BCP Cta Ahorros", 4, "Activo", "Balance", "Activo"],
        ["12.01.01.01", "Facturas por Cobrar Terceros", 4, "Activo", "Balance", "Activo"],
        ["12.01.01.02", "Letras por Cobrar", 4, "Activo", "Balance", "Activo"],
        ["14.01.01.01", "Mercaderias Manufacturadas", 4, "Activo", "Balance", "Activo"],
        ["14.01.02.01", "Materias Primas", 4, "Activo", "Balance", "Activo"],
        ["20.01.01.01", "Inmuebles", 4, "Activo", "Balance", "Activo"],
        ["20.01.02.01", "Maquinaria y Equipo", 4, "Activo", "Balance", "Activo"],
        ["30.01.01.01", "Depreciacion Acumulada", 4, "Activo", "Balance", "Pasivo"],
        ["40.01.01.01", "Tributos por Pagar", 4, "Pasivo", "Balance", "Pasivo"],
        ["40.01.02.01", "Remuneraciones por Pagar", 4, "Pasivo", "Balance", "Pasivo"],
        ["40.01.03.01", "Proveedores Diversos", 4, "Pasivo", "Balance", "Pasivo"],
        ["42.01.01.01", "Prestamos Bancarios CP", 4, "Pasivo", "Balance", "Pasivo"],
        ["50.01.01.01", "Capital Social", 4, "Patrimonio", "Balance", "Patrimonio"],
        ["50.01.02.01", "Reserva Legal", 4, "Patrimonio", "Balance", "Patrimonio"],
        ["50.01.03.01", "Resultados Acumulados", 4, "Patrimonio", "Balance", "Patrimonio"],
        ["60.01.01.01", "Ventas Mercaderias", 4, "Ingreso", "PyG", "Ingreso"],
        ["60.01.02.01", "Prestacion Servicios", 4, "Ingreso", "PyG", "Ingreso"],
        ["70.01.01.01", "Compras Mercaderias", 4, "Gasto", "PyG", "Gasto"],
        ["70.01.02.01", "Servicios Publicos", 4, "Gasto", "PyG", "Gasto"],
        ["80.01.01.01", "Sueldos y Salarios", 4, "Gasto", "PyG", "Gasto"],
        ["80.01.02.01", "Alquileres", 4, "Gasto", "PyG", "Gasto"],
        ["80.01.03.01", "Depreciacion", 4, "Gasto", "PyG", "Gasto"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 6)
    add_border_range(ws, 1, len(data) + 1, 1, 6)
    # validación tipo
    dv = DataValidation(type="list", formula1='"Activo,Pasivo,Patrimonio,Ingreso,Gasto"', allow_blank=True)
    dv.error = "Tipo de cuenta inválido"
    dv.errorTitle = "Error"
    ws.add_data_validation(dv)
    dv.add("D2:D26")
    # formato condicional por tipo
    ws.conditional_formatting.add("A2:F26",
        ColorScaleRule(start_type="min", start_color="C8E6C9",
                       mid_type="percentile", mid_value=50, mid_color="FFF9C4",
                       end_type="max", end_color="FFCDD2"))
    auto_width(ws)


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 3 — Libro Diario y Mayor
# ═══════════════════════════════════════════════════════════════════
def cap3(wb):
    ws = wb.active
    ws.title = "Libro Diario"
    ws["A1"] = "Fecha"; ws["B1"] = "Comprobante"; ws["C1"] = "Glosa"
    ws["D1"] = "Cuenta Debe"; ws["E1"] = "Cuenta Haber"; ws["F1"] = "Debe S/"
    ws["G1"] = "Haber S/"
    data = [
        ["02/01/2025", "CP-001", "Apertura ejercicio", "10.01.01.02", "50.01.01.01", 500000, 500000],
        ["02/01/2025", "CP-001", "Apertura ejercicio", "10.01.02.01", "42.01.01.01", 200000, 200000],
        ["05/01/2025", "CP-002", "Compra mercaderias", "70.01.01.01", "10.01.02.01", 45000, 45000],
        ["08/01/2025", "CP-003", "Venta al contado", "10.01.02.01", "60.01.01.01", 85000, 85000],
        ["10/01/2025", "CP-004", "Pago alquiler enero", "80.01.02.01", "10.01.02.01", 5000, 5000],
        ["12/01/2025", "CP-005", "Cobranza factura", "10.01.02.01", "12.01.01.01", 12000, 12000],
        ["15/01/2025", "CP-006", "Pago sueldos quincena", "80.01.01.01", "10.01.02.01", 25000, 25000],
        ["18/01/2025", "CP-007", "Venta al credito", "12.01.01.01", "60.01.01.01", 32000, 32000],
        ["20/01/2025", "CP-008", "Pago servicios publicos", "70.01.02.01", "10.01.02.01", 1800, 1800],
        ["22/01/2025", "CP-009", "Depreciacion mensual", "80.01.03.01", "30.01.01.01", 4200, 4200],
        ["25/01/2025", "CP-010", "Compra maquinaria", "20.01.02.01", "40.01.03.01", 35000, 35000],
        ["28/01/2025", "CP-011", "Pago tributos", "40.01.01.01", "10.01.02.01", 8500, 8500],
        ["30/01/2025", "CP-012", "Varias ventas contado", "10.01.02.01", "60.01.01.01", 28000, 28000],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 7)
    add_border_range(ws, 1, len(data) + 1, 1, 7)
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=6).number_format = '#,##0.00'
        ws.cell(row=r, column=7).number_format = '#,##0.00'
    ws["A15"] = "TOTAL DEBE"; ws["F15"] = "=SUM(F2:F14)"
    ws["F15"].number_format = '#,##0.00'
    ws["G15"] = "=SUM(G2:G14)"
    ws["G15"].number_format = '#,##0.00'
    ws["A16"] = "CUADRA (Debe=Haber)"; ws["B16"] = '=IF(F15=G15,"SI","NO")'
    ws["B16"].font = Font(bold=True, size=12)
    auto_width(ws)

    # Hoja 2: Libro Mayor
    ws2 = wb.create_sheet("Libro Mayor")
    ws2["A1"] = "Cuenta"; ws2["B1"] = "Descripcion"
    ws2["C1"] = "Debe Total"; ws2["D1"] = "Haber Total"
    ws2["E1"] = "Saldo"
    cuentas = [
        ("10.01.01.02", "Caja General Soles"),
        ("10.01.02.01", "Banco BCP Cta Ahorros"),
        ("12.01.01.01", "Facturas por Cobrar"),
        ("20.01.02.01", "Maquinaria y Equipo"),
        ("30.01.01.01", "Depreciacion Acumulada"),
        ("40.01.01.01", "Tributos por Pagar"),
        ("40.01.03.01", "Proveedores Diversos"),
        ("42.01.01.01", "Prestamos Bancarios CP"),
        ("50.01.01.01", "Capital Social"),
        ("60.01.01.01", "Ventas Mercaderias"),
        ("70.01.01.01", "Compras Mercaderias"),
        ("70.01.02.01", "Servicios Publicos"),
        ("80.01.01.01", "Sueldos y Salarios"),
        ("80.01.02.01", "Alquileres"),
        ("80.01.03.01", "Depreciacion"),
    ]
    diario_debe_col = "F"
    diario_haber_col = "G"
    diario_cuenta_debe = "D"
    diario_cuenta_haber = "E"
    max_row = len(data) + 1
    for i, (codigo, desc) in enumerate(cuentas, 2):
        ws2.cell(row=i, column=1, value=codigo)
        ws2.cell(row=i, column=2, value=desc)
        ws2.cell(row=i, column=3).value = f"=SUMIF({diario_cuenta_debe}2:{diario_cuenta_debe}{max_row},A{i},{diario_debe_col}2:{diario_debe_col}{max_row})"
        ws2.cell(row=i, column=4).value = f"=SUMIF({diario_cuenta_haber}2:{diario_cuenta_haber}{max_row},A{i},{diario_haber_col}2:{diario_haber_col}{max_row})"
        ws2.cell(row=i, column=5).value = f"=C{i}-D{i}"
        for c in range(3, 6):
            ws2.cell(row=i, column=c).number_format = '#,##0.00'
    style_header(ws2, 1, 5)
    add_border_range(ws2, 1, len(cuentas) + 1, 1, 5)
    auto_width(ws2)


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 4 — Funciones Financieras
# ═══════════════════════════════════════════════════════════════════
def cap4(wb):
    ws = wb.active
    ws.title = "Funciones Financieras"
    ws["A1"] = "Proyecto"; ws["B1"] = "Inversion"; ws["C1"] = "Flujo Anual"
    ws["D1"] = "Tasa Descuento"; ws["E1"] = "VAN"; ws["F1"] = "TIR"
    ws["G1"] = "Cuota Mensual"
    data = [
        ["Proy A - Ampliacion Planta", 500000, 150000, 0.10, None, None, None],
        ["Proy B - Nuevo Producto", 300000, 85000, 0.10, None, None, None],
        ["Proy C - Automatizacion", 200000, 55000, 0.10, None, None, None],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    # TIR assume 5-year project
    for i in range(2, 5):
        ws.cell(row=i, column=7).value = f"=PMT(15%/12,36,B{i})"
        ws.cell(row=i, column=7).number_format = '#,##0.00'
    style_header(ws, 1, 7)
    add_border_range(ws, 1, 4, 1, 7)

    # Tabla prestamo
    ws["A7"] = "TABLA AMORTIZACION"; ws["A7"].font = Font(bold=True, size=12, color="0A7E5C")
    ws["A8"] = "Monto Prestamo"; ws["B8"] = 150000
    ws["B8"].number_format = '#,##0.00'
    ws["A9"] = "Tasa Anual"; ws["B9"] = 0.12
    ws["B9"].number_format = '0.00%'
    ws["A10"] = "Plazo (meses)"; ws["B10"] = 36
    ws["A11"] = "Cuota Mensual"; ws["B11"] = "=PMT(B9/12,B10,B8)"
    ws["B11"].number_format = '#,##0.00'
    ws["A13"] = "Periodo"; ws["B13"] = "Saldo Inicial"; ws["C13"] = "Interes"
    ws["D13"] = "Amortizacion"; ws["E13"] = "Cuota"; ws["F13"] = "Saldo Final"
    style_header(ws, 13, 6)
    tasa_mensual = 0.12 / 12
    cuota = 150000 * tasa_mensual * (1 + tasa_mensual) ** 36 / ((1 + tasa_mensual) ** 36 - 1)
    saldo = 150000.0
    for per in range(1, 13):
        r = per + 13
        interes = saldo * tasa_mensual
        amort = cuota - interes
        saldo_final = saldo - amort
        ws.cell(row=r, column=1, value=per)
        ws.cell(row=r, column=2, value=round(saldo, 2))
        ws.cell(row=r, column=3, value=round(interes, 2))
        ws.cell(row=r, column=4, value=round(amort, 2))
        ws.cell(row=r, column=5, value=round(cuota, 2))
        ws.cell(row=r, column=6, value=round(saldo_final, 2))
        for c in range(2, 7):
            ws.cell(row=r, column=c).number_format = '#,##0.00'
        saldo = saldo_final
    add_border_range(ws, 13, 25, 1, 6)
    auto_width(ws)


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 5 — Funciones Condicionales
# ═══════════════════════════════════════════════════════════════════
def cap5(wb):
    ws = wb.active
    ws.title = "Gastos Mensuales"
    ws["A1"] = "Fecha"; ws["B1"] = "Proveedor"; ws["C1"] = "Categoria"
    ws["D1"] = "Centro Costo"; ws["E1"] = "Monto"; ws["F1"] = "Clasificacion"
    ws["G1"] = "Presupuesto"
    cats = [
        ("Servicios", "Gasto Operativo", 50000),
        ("Alquiler", "Gasto Operativo", 6000),
        ("Planilla", "Gasto Administrativo", 85000),
        ("Marketing", "Gasto Administrativo", 12000),
        ("Materia Prima", "Costo Venta", 60000),
        ("Transporte", "Costo Venta", 15000),
        ("Seguros", "Gasto Operativo", 3000),
        ("Mantenimiento", "Gasto Operativo", 4500),
        ("Impuestos", "Gasto Administrativo", 18000),
        ("Comisiones", "Costo Venta", 8000),
    ]
    proveedores = ["Maderas Perez", "Pinturas Unidas", "Luz del Sur", "Sedapal",
                   "Planilla", "Agencia Creativa", "Logistica Total",
                   "Pacifico Seguros", "SUNAT", "Interbank"]
    cc = ["CC-ADM", "CC-VTA", "CC-PROD", "CC-ADM", "CC-PROD", "CC-VTA",
          "CC-ADM", "CC-PROD", "CC-ADM", "CC-VTA"]
    random.seed(123)
    for i in range(1, 101):
        r = i + 1
        mes = random.randint(1, 6)
        dia = random.randint(1, 28)
        ws.cell(row=r, column=1, value=f"{dia:02d}/{mes:02d}/2025")
        idx = random.randint(0, 9)
        ws.cell(row=r, column=2, value=proveedores[idx])
        cat = cats[idx][0]
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=cc[idx])
        monto = round(random.uniform(500, 12000), 2)
        ws.cell(row=r, column=5, value=monto)
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        # Clasificacion con IFS
        ws.cell(row=r, column=6).value = f'=IFS(C{r}="Materia Prima","Costo Venta",C{r}="Transporte","Costo Venta",C{r}="Comisiones","Costo Venta",C{r}="Planilla","Gasto Administrativo",C{r}="Marketing","Gasto Administrativo",C{r}="Impuestos","Gasto Administrativo",TRUE,"Gasto Operativo")'
        ws.cell(row=r, column=7, value=cats[idx][2])
        ws.cell(row=r, column=7).number_format = '#,##0.00'
    style_header(ws, 1, 7)
    add_border_range(ws, 1, 101, 1, 7)
    # resumen
    ws["A103"] = "RESUMEN POR CENTRO COSTO"
    ws["A103"].font = Font(bold=True, size=12, color="0A7E5C")
    ws["A104"] = "Centro Costo"; ws["B104"] = "Total Gastos"
    style_header(ws, 104, 2)
    cc_unique = sorted(set(cc))
    for i, c in enumerate(cc_unique):
        r = 105 + i
        ws.cell(row=r, column=1, value=c)
        ws.cell(row=r, column=2).value = f"=SUMIF(D2:D101,A{r},E2:E101)"
        ws.cell(row=r, column=2).number_format = '#,##0.00'
    add_border_range(ws, 104, 108, 1, 2)
    auto_width(ws)


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 6 — Tablas Dinámicas
# ═══════════════════════════════════════════════════════════════════
def cap6(wb):
    ws = wb.active
    ws.title = "Ventas Anuales"
    ws["A1"] = "Fecha"; ws["B1"] = "Producto"; ws["C1"] = "Categoria"
    ws["D1"] = "Region"; ws["E1"] = "Cantidad"; ws["F1"] = "Precio Unit."
    ws["G1"] = "Total Venta"
    productos = [
        ("Silla Ejecutiva", "Muebles"), ("Mesa Directiva", "Muebles"),
        ("Estante Archivo", "Muebles"), ("Escritorio Gerencial", "Muebles"),
        ("Laptop Contabilidad", "Equipos"), ("Monitor 24", "Equipos"),
        ("Impresora Fiscal", "Equipos"), ("Papel Bond", "Suministros"),
        ("Tinta Impresora", "Suministros"), ("Archivador", "Suministros"),
    ]
    regiones = ["Lima", "Arequipa", "Cusco", "Trujillo", "Chiclayo"]
    random.seed(456)
    for i in range(1, 201):
        r = i + 1
        mes = random.randint(1, 12)
        dia = random.randint(1, 28)
        ws.cell(row=r, column=1, value=f"{dia:02d}/{mes:02d}/2025")
        prod = random.choice(productos)
        ws.cell(row=r, column=2, value=prod[0])
        ws.cell(row=r, column=3, value=prod[1])
        ws.cell(row=r, column=4, value=random.choice(regiones))
        cant = random.randint(1, 30)
        ws.cell(row=r, column=5, value=cant)
        precio = round(random.uniform(15, 450), 2)
        ws.cell(row=r, column=6, value=precio)
        ws.cell(row=r, column=6).number_format = '#,##0.00'
        ws.cell(row=r, column=7).value = f"=E{r}*F{r}"
        ws.cell(row=r, column=7).number_format = '#,##0.00'
    style_header(ws, 1, 7)
    add_border_range(ws, 1, 201, 1, 7)
    auto_width(ws)


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 7 — Consolidación
# ═══════════════════════════════════════════════════════════════════
def cap7(wb):
    ws = wb.active
    ws.title = "Sucursal Lima"
    ws["A1"] = "Cuenta"; ws["B1"] = "Enero"; ws["C1"] = "Febrero"
    ws["D1"] = "Marzo"; ws["E1"] = "Abril"; ws["F1"] = "Total"
    data = [
        ["Ventas", 120000, 135000, 128000, 142000, None],
        ["Costo Ventas", -72000, -81000, -76800, -85200, None],
        ["Gastos Operativos", -25000, -26000, -24800, -27000, None],
        ["Gastos Admin", -12000, -12500, -11800, -13000, None],
        ["Utilidad Operativa", None, None, None, None, None],
        ["Gastos Financieros", -3000, -3000, -3200, -3000, None],
        ["Utilidad Neta", None, None, None, None, None],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    for c in range(2, 7):
        col_letter = get_column_letter(c)
        r5 = 2  # Ventas row
        r6 = 3  # Costo Ventas
        r7 = 4  # Gastos Operativos
        r8 = 5  # Gastos Admin
        ws.cell(row=6, column=c).value = f"={col_letter}2+{col_letter}3+{col_letter}4+{col_letter}5"
        ws.cell(row=8, column=c).value = f"={col_letter}6+{col_letter}7"
    style_header(ws, 1, 6)
    add_border_range(ws, 1, 8, 1, 6)
    for r in range(2, 9):
        for c in range(2, 7):
            ws.cell(row=r, column=c).number_format = '#,##0.00'
    # Totales column
    for r in range(2, 9):
        ws.cell(row=r, column=6).value = f"=SUM(B{r}:E{r})"
        ws.cell(row=r, column=6).number_format = '#,##0.00'
    auto_width(ws)


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 8 — Auditoría
# ═══════════════════════════════════════════════════════════════════
def cap8(wb):
    ws = wb.active
    ws.title = "Auditoria"
    ws["A1"] = "Transaccion"; ws["B1"] = "Cuenta"; ws["C1"] = "Monto"
    ws["D1"] = "Fecha"; ws["E1"] = "Centro Costo"; ws["F1"] = "Auditado"
    ws["G1"] = "Observacion"
    random.seed(789)
    for i in range(1, 101):
        r = i + 1
        ws.cell(row=r, column=1, value=f"TXN-{i:04d}")
        ws.cell(row=r, column=2, value=f"10.01.{random.randint(1,99):02d}.{random.randint(1,99):02d}")
        monto = round(random.uniform(100, 50000), 2)
        if random.random() < 0.05:  # 5% errores
            if random.random() < 0.5:
                ws.cell(row=r, column=3, value="ERROR")
            else:
                ws.cell(row=r, column=3, value=-monto)  # monto negativo sin sentido
        else:
            ws.cell(row=r, column=3, value=monto)
        ws.cell(row=r, column=4, value=f"{random.randint(1,28):02d}/{random.randint(1,12):02d}/2025")
        ws.cell(row=r, column=5, value=random.choice(["CC-ADM", "CC-VTA", "CC-PROD"]))
        ws.cell(row=r, column=6, value="PENDIENTE")
        ws.cell(row=r, column=7).value = ""  # placeholder
    style_header(ws, 1, 7)
    add_border_range(ws, 1, 101, 1, 7)
    # Valores duplicados de muestra
    for r in [5, 10, 15]:
        ws.cell(row=r, column=3, value=ws.cell(row=3, column=3).value)
    ws.conditional_formatting.add("C2:C101",
        IconSetRule(icon_style="3TrafficLights1", type="percent", values=[0, 33, 67], showValue=True))
    ws.conditional_formatting.add("A2:F101",
        ColorScaleRule(start_type="min", start_color="FFCDD2",
                       mid_type="percentile", mid_value=50, mid_color="FFF9C4",
                       end_type="max", end_color="C8E6C9"))
    auto_width(ws)

    # Hoja extracto bancario
    ws2 = wb.create_sheet("Extracto Bancario")
    ws2["A1"] = "Fecha"; ws2["B1"] = "Referencia"; ws2["C1"] = "Monto"
    ws2["D1"] = "Tipo"
    random.seed(111)
    for i in range(1, 51):
        r = i + 1
        ws2.cell(row=r, column=1, value=f"{random.randint(1,28):02d}/{random.randint(1,12):02d}/2025")
        ws2.cell(row=r, column=2, value=f"BCO-{i:04d}")
        ws2.cell(row=r, column=3, value=round(random.uniform(500, 30000), 2))
        ws2.cell(row=r, column=3).number_format = '#,##0.00'
        ws2.cell(row=r, column=4, value=random.choice(["DEPOSITO", "RETIRO", "TRANSFERENCIA"]))
    style_header(ws2, 1, 4)
    add_border_range(ws2, 1, 51, 1, 4)
    auto_width(ws2)


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 9 — Dashboard
# ═══════════════════════════════════════════════════════════════════
def cap9(wb):
    ws = wb.active
    ws.title = "Datos Dashboard"
    # Datos mensuales
    ws["A1"] = "Mes"; ws["B1"] = "Ventas"; ws["C1"] = "Gastos"
    ws["D1"] = "Utilidad"; ws["E1"] = "Liquidez"; ws["F1"] = "ROE"
    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    ventas = [180000, 195000, 172000, 210000, 225000, 198000,
              240000, 255000, 230000, 270000, 290000, 310000]
    gastos = [135000, 142000, 128000, 155000, 160000, 148000,
              172000, 180000, 165000, 190000, 195000, 205000]
    for i, m in enumerate(meses, 2):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=ventas[i - 2])
        ws.cell(row=i, column=3, value=gastos[i - 2])
        ws.cell(row=i, column=4).value = f"=B{i}-C{i}"
        ws.cell(row=i, column=5, value=round(random.uniform(1.2, 2.5), 2))
        ws.cell(row=i, column=6, value=round(random.uniform(8.0, 18.0), 1))
        for c in range(2, 5):
            ws.cell(row=i, column=c).number_format = '#,##0'
    style_header(ws, 1, 6)
    add_border_range(ws, 1, 13, 1, 6)
    # Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Ventas vs Gastos 2025"
    chart.style = 10
    chart.y_axis.title = "Soles"
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=13)
    cats = Reference(ws, min_col=1, min_row=2, max_row=13)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 22
    chart.height = 14
    ws.add_chart(chart, "A15")
    auto_width(ws)

    # Hoja KPIs
    ws2 = wb.create_sheet("KPIs")
    ws2["A1"] = "KPI"; ws2["B1"] = "Valor Actual"; ws2["C1"] = "Meta"
    ws2["D1"] = "Cumplimiento"
    kpis = [
        ["Liquidez Corriente", 2.15, 1.5, None],
        ["Prueba Acida", 1.42, 1.0, None],
        ["ROE (%)", 15.3, 12.0, None],
        ["ROA (%)", 8.7, 7.0, None],
        ["Margen Neto (%)", 22.4, 18.0, None],
        ["Endeudamiento", 0.45, 0.6, None],
        ["Cobertura Intereses", 4.8, 3.0, None],
    ]
    for i, row in enumerate(kpis, 2):
        ws2.cell(row=i, column=1, value=row[0])
        ws2.cell(row=i, column=2, value=row[1])
        ws2.cell(row=i, column=3, value=row[2])
        ws2.cell(row=i, column=4).value = f"=B{i}/C{i}"
        ws2.cell(row=i, column=4).number_format = '0.00%'
    style_header(ws2, 1, 4)
    add_border_range(ws2, 1, len(kpis) + 1, 1, 4)
    ws2.conditional_formatting.add("B2:B8",
        DataBarRule(start_type="min", end_type="max", color="0A7E5C"))
    auto_width(ws2)


# ═══════════════════════════════════════════════════════════════════
#  CAPÍTULO 10 — Macros VBA
# ═══════════════════════════════════════════════════════════════════
def cap10(wb):
    ws = wb.active
    ws.title = "Conciliacion Macro"
    ws["A1"] = "Fecha"; ws["B1"] = "Referencia"; ws["C1"] = "Descripcion"
    ws["D1"] = "Monto S/"; ws["E1"] = "Estado"; ws["F1"] = "Diferencia"
    random.seed(999)
    for i in range(1, 51):
        r = i + 1
        ws.cell(row=r, column=1, value=f"{random.randint(1,28):02d}/{random.randint(1,12):02d}/2025")
        ws.cell(row=r, column=2, value=f"REF-{i:04d}")
        ws.cell(row=r, column=3, value=random.choice(["Pago proveedor", "Cobranza cliente", "Compra materiales",
                                                       "Venta contado", "Pago servicios", "Deposito bancario"]))
        ws.cell(row=r, column=4, value=round(random.uniform(1000, 50000), 2))
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        ws.cell(row=r, column=5, value=random.choice(["PENDIENTE", "CONCILIADO"]))
        ws.cell(row=r, column=6).value = f"=SI.ERROR(D{r}-BUSCARX(B{r},'Extracto Bancario'!B:B,'Extracto Bancario'!C:C),D{r})"
        ws.cell(row=r, column=6).number_format = '#,##0.00'
    style_header(ws, 1, 6)
    add_border_range(ws, 1, 51, 1, 6)
    ws.conditional_formatting.add("E2:E51",
        IconSetRule(icon_style="3TrafficLights1", type="formula", values=[0, 33, 67], showValue=True))
    auto_width(ws)

    # Hoja extracto
    ws2 = wb.create_sheet("Extracto Bancario")
    ws2["A1"] = "Referencia"; ws2["B1"] = "Monto"
    refs = [f"REF-{i:04d}" for i in range(1, 46)]  # 5 no conciliadas
    random.shuffle(refs)
    for i, ref in enumerate(refs, 2):
        ws2.cell(row=i, column=1, value=ref)
        ws2.cell(row=i, column=2, value=round(random.uniform(1000, 50000), 2))
        ws2.cell(row=i, column=2).number_format = '#,##0.00'
    style_header(ws2, 1, 2)
    add_border_range(ws2, 1, len(refs) + 1, 1, 2)
    auto_width(ws2)


# ═══════════════════════════════════════════════════════════════════
#  GENERACIÓN
# ═══════════════════════════════════════════════════════════════════
CHAPTERS = [
    ("c01_configuracion", "Configuracion y Conciliacion", cap1),
    ("c02_catalogo", "Catalogo de Cuentas", cap2),
    ("c03_diario_mayor", "Libro Diario y Mayor", cap3),
    ("c04_financieras", "Funciones Financieras", cap4),
    ("c05_condicionales", "Funciones Condicionales", cap5),
    ("c06_tablas_dinamicas", "Tablas Dinamicas", cap6),
    ("c07_consolidacion", "Consolidacion", cap7),
    ("c08_auditoria", "Auditoria y Validacion", cap8),
    ("c09_dashboard", "Dashboard y KPIs", cap9),
    ("c10_vba", "Macros VBA", cap10),
]

def main():
    DEST.mkdir(parents=True, exist_ok=True)
    for nombre, titulo, fn in CHAPTERS:
        wb = Workbook()
        fn(wb)
        ruta = DEST / f"{nombre}.xlsx"
        wb.save(ruta)
        print(f"  OK {nombre}.xlsx - {titulo}")
    print(f"\nOK - Todos los archivos en:\n   {DEST}")

if __name__ == "__main__":
    main()
