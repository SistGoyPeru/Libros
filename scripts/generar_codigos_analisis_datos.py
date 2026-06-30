"""
Genera los archivos de ejemplo (codigos/) para Excel para Análisis de Datos
"""

import os, csv, random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CODIGOS = os.path.join(BASE, "excel_analisis_datos", "codigos")
random.seed(42)

HEADER_FILL = PatternFill("solid", fgColor="0A7E5C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def stylize_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

def center_cols(ws, cols):
    for col in cols:
        for cell in ws[col]:
            cell.alignment = Alignment(horizontal="center")

def add_borders(ws, max_row, max_col):
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER

# ── c01_proceso_analisis ───────────────────────────────────────────────
def gen_c01():
    out = os.path.join(CODIGOS, "c01_proceso_analisis")
    os.makedirs(out, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Proceso Analisis"
    ws.append(["Etapa", "Actividad", "Descripcion", "Herramientas", "DuracionHoras"])
    etapas = [
        ("1. Definir Problema", "Identificar objetivos", "Definir preguntas de negocio e hipotesis", "Reunion stakeholders", 4),
        ("2. Recopilar Datos", "Obtener fuentes", "Identificar y extraer datos de fuentes internas/externas", "SQL, Excel, APIs", 8),
        ("3. Limpiar Datos", "Preparacion inicial", "Manejar nulos, duplicados, formatos inconsistentes", "Excel, Python", 12),
        ("4. Explorar Datos", "Analisis exploratorio", "Estadisticas descriptivas, visualizaciones iniciales", "Excel, Power BI", 6),
        ("5. Analizar Datos", "Modelado y analisis", "Aplicar tecnicas estadisticas y modelos", "Excel, Python, R", 16),
        ("6. Interpretar", "Conclusiones", "Traducir resultados a recomendaciones de negocio", "Presentacion", 4),
        ("7. Comunicar", "Dashboard/Reporte", "Crear visualizaciones y presentar hallazgos", "Excel, Power BI", 6),
        ("8. Implementar", "Accion", "Ejecutar recomendaciones y medir impacto", "Seguimiento", 4),
    ]
    for e in etapas:
        ws.append(e)
    stylize_header(ws, 5)
    center_cols(ws, ["E"])
    add_borders(ws, ws.max_row, 5)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 16
    wb.save(os.path.join(out, "proceso_analisis.xlsx"))
    print(f"  OK: c01_proceso_analisis/proceso_analisis.xlsx")

    # Preguntas de negocio - CSV
    with open(os.path.join(out, "preguntas_negocio.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ID", "Pregunta", "Tipo", "Prioridad", "FuenteDatos"])
        preguntas = [
            (1, "Cuales son los productos mas vendidos?", "Descriptivo", "Alta", "Ventas"),
            (2, "Como varian las ventas por temporada?", "Diagnostico", "Alta", "Ventas historicas"),
            (3, "Que factores influyen en la satisfaccion del cliente?", "Diagnostico", "Media", "Encuestas"),
            (4, "Cual sera la demanda del proximo trimestre?", "Predictivo", "Alta", "Ventas historicas"),
            (5, "Que segmento de clientes es mas rentable?", "Descriptivo", "Media", "Clientes"),
            (6, "Como optimizar el inventario?", "Prescriptivo", "Baja", "Inventario"),
            (7, "Cual es el ticket promedio por region?", "Descriptivo", "Media", "Ventas"),
            (8, "Existe correlacion entre gasto en marketing y ventas?", "Diagnostico", "Alta", "Marketing, Ventas"),
        ]
        for p in preguntas:
            w.writerow(p)
    print(f"  OK: c01_proceso_analisis/preguntas_negocio.csv")


# ── c02_limpieza_datos ─────────────────────────────────────────────────
def gen_c02():
    out = os.path.join(CODIGOS, "c02_limpieza_datos")
    os.makedirs(out, exist_ok=True)

    # Datos sucios - CSV
    with open(os.path.join(out, "datos_sucios.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ID", "Nombre", "Edad", "Email", "Telefono", "Monto", "FechaIngreso", "Region"])
        dirty_data = [
            (1, "Juan Perez", 30, "juan@email.com", "999-123-456", 1500.50, "2025-01-15", "Lima"),
            (2, "Maria Lopez", None, "maria@email.com", "998-654-321", 0, "2025-01-20", "Arequipa"),
            (3, "Carlos Ruiz", -5, "carlos@", "997-111-222", 2500.00, "2025/01/25", "cusco"),
            (4, "Ana Torres", 28, "ana.torres@email.com", None, 3200.00, "2025-02-01", "Lima"),
            (5, "Pedro Soto", 45, "pedro@email.com", "996-333-444", -500.00, "2025-02-10", "Trujillo"),
            (6, "Luisa Vega", 33, None, "995-555-666", 1800.00, "2025-02-15", "Lima"),
            (7, "Mario Diaz", 22, "mario@", "994-777-888", 0, "15/03/2025", "arequipa"),
            (8, "Sofia Rios", 29, "sofia@email.com", "993-111-999", 2100.00, "2025-03-01", "Cusco"),
            (9, "Diego Paz", None, "diego@email.com", None, 0, "2025-03-10", "Lima"),
            (10, "Claudia Sol", 31, "claudia@email.com", "992-222-333", 1750.00, "2025-03-15", "Piura"),
            (11, "Jose Ramos", -1, "jose@email.com", "991-444-555", 3000.00, None, "Lima"),
            (12, "Rosa Cruz", 40, "rosa@email.com", "990-666-777", 2200.00, "2025-04-01", ""),
            (13, "Luis Gil", 52, "luis.gil@email.com", None, 0, "2025-04-10", "Lima"),
        ]
        for d in dirty_data:
            w.writerow(d)
    print(f"  OK: c02_limpieza_datos/datos_sucios.csv")

    # Guia de limpieza - xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Guia Limpieza"
    ws.append(["Problema", "Descripcion", "Metodo de Limpieza", "Formula/Herramienta"])
    problemas = [
        ("Valores Nulos", "Celdas vacias en Edad, Email, Telefono", "Imputar con media, mediana o eliminar filas", "=SI.ERROR(A2, \"\")"),
        ("Edades Negativas", "Edades con valores -5, -1", "Reemplazar por valor valido o eliminar", "=SI(A2>0, A2, \"\")"),
        ("Formato Fecha", "Fechas en formatos inconsistentes", "Estandarizar a YYYY-MM-DD", "=TEXTO(A2, \"aaaa-mm-dd\")"),
        ("Regiones Inconsistentes", "cusco vs Cusco, arequipa vs Arequipa", "Normalizar a mayuscula inicial", "=MAYUSC.INIC(A2)"),
        ("Monto Cero", "Registros con monto = 0", "Verificar si es real o error de carga", "Filtro > 0"),
        ("Email Invalido", "Emails sin @ o dominio", "Validar con mascara", "=SI(CONTAR.SI(A2, \"*@*.*\")>0, \"OK\", \"Invalido\")"),
        ("Duplicados", "Registros repetidos", "Eliminar duplicados", "Quitar duplicados en Datos"),
        ("Espacios Extras", "Espacios al inicio/final del texto", "Limpiar con funcion", "=ESPACIOS(A2)"),
    ]
    for p in problemas:
        ws.append(p)
    stylize_header(ws, 4)
    add_borders(ws, ws.max_row, 4)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 36
    wb.save(os.path.join(out, "guia_limpieza.xlsx"))
    print(f"  OK: c02_limpieza_datos/guia_limpieza.xlsx")


# ── c03_eda ─────────────────────────────────────────────────────────────
def gen_c03():
    out = os.path.join(CODIGOS, "c03_eda")
    os.makedirs(out, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas EDA"
    ws.append(["Fecha", "Producto", "Categoria", "Cantidad", "PrecioUnitario", "Monto", "Cliente", "Region"])
    productos = [
        ("Silla Ejecutiva", "Muebles", 350), ("Escritorio", "Muebles", 850),
        ("Laptop Pro", "Electrónicos", 3200), ("Monitor 27\"", "Electrónicos", 1200),
        ("Teclado", "Electrónicos", 250), ("Mouse", "Electrónicos", 180),
        ("Archivador", "Oficina", 65), ("Papel Bond", "Oficina", 25),
    ]
    regiones = ["Lima", "Arequipa", "Cusco", "Trujillo", "Piura"]
    clientes = ["CorpAlpha", "BetaSol", "GammaCorp", "DeltaEnt", "Epsilon"]

    for _ in range(500):
        prod, cat, pu = random.choice(productos)
        cant = random.randint(1, 50)
        dia = random.randint(1, 28)
        mes = random.randint(1, 12)
        ws.append([
            f"2025-{mes:02d}-{dia:02d}", prod, cat, cant, pu,
            round(cant * pu, 2), random.choice(clientes), random.choice(regiones),
        ])
    stylize_header(ws, 8)
    add_borders(ws, ws.max_row, 8)
    ws.freeze_panes = "A2"

    # Resumen estadistico
    ws2 = wb.create_sheet("Resumen Estadistico")
    ws2.append(["Medida", "Cantidad", "PrecioUnitario", "Monto"])
    ws2.append(["Media", "=PROMEDIO(Ventas EDA!D:D)", "=PROMEDIO(Ventas EDA!E:E)", "=PROMEDIO(Ventas EDA!F:F)"])
    ws2.append(["Mediana", "=MEDIANA(Ventas EDA!D:D)", "=MEDIANA(Ventas EDA!E:E)", "=MEDIANA(Ventas EDA!F:F)"])
    ws2.append(["Desv Estandar", "=DESVEST(Ventas EDA!D:D)", "=DESVEST(Ventas EDA!E:E)", "=DESVEST(Ventas EDA!F:F)"])
    ws2.append(["Minimo", "=MIN(Ventas EDA!D:D)", "=MIN(Ventas EDA!E:E)", "=MIN(Ventas EDA!F:F)"])
    ws2.append(["Maximo", "=MAX(Ventas EDA!D:D)", "=MAX(Ventas EDA!E:E)", "=MAX(Ventas EDA!F:F)"])
    ws2.append(["Conteo", "=CONTARA(Ventas EDA!D:D)", "=CONTARA(Ventas EDA!E:E)", "=CONTARA(Ventas EDA!F:F)"])
    stylize_header(ws2, 4)
    add_borders(ws2, ws2.max_row, 4)

    # Tabla dinamica resumen por producto
    ws3 = wb.create_sheet("Resumen por Producto")
    ws3.append(["Producto", "Total Cantidad", "Total Monto", "Promedio Monto", "Conteo"])
    products_uniq = ["Silla Ejecutiva", "Escritorio", "Laptop Pro", "Monitor 27\"", "Teclado", "Mouse", "Archivador", "Papel Bond"]
    for p in products_uniq:
        ws3.append([p, 0, 0, 0, 0])
    stylize_header(ws3, 5)
    add_borders(ws3, ws3.max_row, 5)

    wb.save(os.path.join(out, "eda_ventas.xlsx"))
    print(f"  OK: c03_eda/eda_ventas.xlsx")

    # CSV con datos adicionales para EDA
    with open(os.path.join(out, "clientes_segmentos.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ID_Cliente", "Nombre", "Segmento", "IngresoAnual", "FrecuenciaCompra", "AntiguedadMeses"])
        for i in range(1, 51):
            w.writerow([
                i,
                f"Cliente_{i:03d}",
                random.choice(["VIP", "Premium", "Regular", "Basico"]),
                round(random.uniform(20000, 200000), 2),
                random.randint(1, 24),
                random.randint(1, 72),
            ])
    print(f"  OK: c03_eda/clientes_segmentos.csv")


# ── c04_funciones ──────────────────────────────────────────────────────
def gen_c04():
    out = os.path.join(CODIGOS, "c04_funciones")
    os.makedirs(out, exist_ok=True)

    wb = Workbook()
    # Ventas table
    ws1 = wb.active
    ws1.title = "Ventas"
    ws1.append(["ID_Venta", "Fecha", "ID_Producto", "ID_Cliente", "Cantidad", "PrecioUnitario", "Monto"])
    productos_venta = [101, 102, 103, 104, 105, 106, 107, 108]
    clientes_venta = [201, 202, 203, 204, 205]
    for i in range(1, 201):
        pid = random.choice(productos_venta)
        cant = random.randint(1, 20)
        pu = random.uniform(10, 500)
        ws1.append([
            f"V{i:04d}",
            f"2025-{random.randint(1,12):02d}-{random.randint(1,28):02d}",
            pid,
            random.choice(clientes_venta),
            cant,
            round(pu, 2),
            round(cant * pu, 2),
        ])
    stylize_header(ws1, 7)
    center_cols(ws1, ["A", "C", "D", "E"])
    add_borders(ws1, ws1.max_row, 7)

    ws2 = wb.create_sheet("Productos")
    ws2.append(["ID_Producto", "Nombre", "Categoria", "Precio"])
    prods = [
        (101, "Silla Ejecutiva", "Muebles", 350.00),
        (102, "Escritorio", "Muebles", 850.00),
        (103, "Laptop Pro", "Electrónicos", 3200.00),
        (104, "Monitor 27\"", "Electrónicos", 1200.00),
        (105, "Teclado Mecánico", "Electrónicos", 250.00),
        (106, "Mouse Inalámbrico", "Electrónicos", 180.00),
        (107, "Archivador", "Oficina", 65.00),
        (108, "Papel Bond x500", "Oficina", 25.00),
    ]
    for p in prods:
        ws2.append(p)
    stylize_header(ws2, 4)
    center_cols(ws2, ["A"])
    add_borders(ws2, ws2.max_row, 4)

    ws3 = wb.create_sheet("Clientes")
    ws3.append(["ID_Cliente", "Nombre", "Categoria", "Region"])
    clis = [
        (201, "Corporación Alpha", "VIP", "Lima"),
        (202, "Beta Solutions", "Regular", "Arequipa"),
        (203, "Gamma Trading", "VIP", "Lima"),
        (204, "Delta Import", "Regular", "Cusco"),
        (205, "Epsilon Services", "Premium", "Trujillo"),
    ]
    for c in clis:
        ws3.append(c)
    stylize_header(ws3, 4)
    center_cols(ws3, ["A", "C", "D"])
    add_borders(ws3, ws3.max_row, 4)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
    wb.save(os.path.join(out, "datos_analisis.xlsx"))
    print(f"  OK: c04_funciones/datos_analisis.xlsx")

# ── c05_tendencias ─────────────────────────────────────────────────────
def gen_c05():
    out = os.path.join(CODIGOS, "c05_tendencias")
    os.makedirs(out, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas Historicas"
    ws.append(["Fecha", "Producto", "Categoria", "Cantidad", "Monto", "Region"])
    productos = ["Silla", "Escritorio", "Laptop", "Monitor", "Teclado", "Mouse"]
    categorias = {"Silla": "Muebles", "Escritorio": "Muebles", "Laptop": "Electrónicos",
                  "Monitor": "Electrónicos", "Teclado": "Electrónicos", "Mouse": "Electrónicos"}
    regiones = ["Lima", "Arequipa", "Cusco", "Trujillo", "Piura"]

    for year in [2023, 2024, 2025]:
        for month in range(1, 13):
            # simulate seasonality
            factor = 1.0 + 0.3 * (month in [3, 4, 11, 12]) - 0.2 * (month in [1, 2])
            for prod in productos:
                if random.random() < 0.4:
                    continue
                cant = int(random.randint(5, 50) * factor)
                pu = {"Silla": 350, "Escritorio": 850, "Laptop": 3200,
                      "Monitor": 1200, "Teclado": 250, "Mouse": 180}[prod]
                ws.append([
                    f"{year}-{month:02d}-{random.randint(1,28):02d}",
                    prod,
                    categorias[prod],
                    cant,
                    round(cant * pu * random.uniform(0.9, 1.1), 2),
                    random.choice(regiones),
                ])
    stylize_header(ws, 6)
    center_cols(ws, ["A", "B", "C", "D", "F"])
    add_borders(ws, ws.max_row, 6)
    ws.freeze_panes = "A2"
    wb.save(os.path.join(out, "ventas_historicas.xlsx"))
    print(f"  OK: c05_tendencias/ventas_historicas.xlsx")

# ── c06_tablas_dinamicas ──────────────────────────────────────────────
def gen_c06():
    out = os.path.join(CODIGOS, "c06_tablas_dinamicas")
    os.makedirs(out, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Fecha", "Producto", "Categoria", "Cantidad", "PrecioUnitario", "Monto", "Cliente", "Region", "Segmento"])
    prods = [
        ("Silla Ejecutiva", "Muebles", 350), ("Escritorio", "Muebles", 850),
        ("Laptop Pro", "Electrónicos", 3200), ("Monitor 27\"", "Electrónicos", 1200),
        ("Teclado Mecánico", "Electrónicos", 250), ("Mouse", "Electrónicos", 180),
        ("Archivador", "Oficina", 65), ("Papel Bond", "Oficina", 25),
        ("Impresora", "Electrónicos", 900), ("Webcam HD", "Electrónicos", 150),
    ]
    clientes = [("CorpAlpha", "VIP"), ("BetaSol", "Premium"), ("GammaCorp", "Regular"),
                ("DeltaEnt", "VIP"), ("Epsilon", "Premium"), ("ZetaCorp", "Regular"),
                ("EtaGroup", "VIP"), ("ThetaInc", "Regular")]
    regiones = ["Lima", "Arequipa", "Cusco", "Trujillo", "Piura", "Lambayeque", "Ica", "Junin"]

    for i in range(1, 5001):
        prod, cat, pu = random.choice(prods)
        cant = random.randint(1, 30)
        anno = random.choice([2024, 2025])
        mes = random.randint(1, 12)
        dia = random.randint(1, 28)
        cli, seg = random.choice(clientes)
        reg = random.choice(regiones)
        ws.append([
            f"{anno}-{mes:02d}-{dia:02d}", prod, cat, cant, pu,
            round(cant * pu, 2), cli, reg, seg,
        ])
    stylize_header(ws, 9)
    center_cols(ws, ["A", "B", "C", "D", "E", "G", "H", "I"])
    add_borders(ws, ws.max_row, 9)
    ws.freeze_panes = "A2"
    wb.save(os.path.join(out, "ventas_analisis.xlsx"))
    print(f"  OK: c06_tablas_dinamicas/ventas_analisis.xlsx")

# ── c07_visualizacion ─────────────────────────────────────────────────
def gen_c07():
    out = os.path.join(CODIGOS, "c07_visualizacion")
    os.makedirs(out, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(["Mes", "Ventas", "Gastos", "MetaVentas", "Utilidad", "%Crecimiento"])

    base_ventas = 100000
    base_gastos = 70000
    for m in range(1, 25):
        ventas = int(base_ventas * (1 + m * 0.02 + random.uniform(-0.05, 0.05)))
        gastos = int(base_gastos * (1 + m * 0.01 + random.uniform(-0.03, 0.03)))
        meta = int(base_ventas * 1.1 * (1 + m * 0.015))
        utilidad = ventas - gastos
        crecimiento = (ventas / (base_ventas * (1 + (m - 1) * 0.02)) - 1) if m > 1 else 0
        ws.append([
            f"2025-{m:02d}", ventas, gastos, meta, utilidad,
            round(crecimiento, 4),
        ])
    stylize_header(ws, 6)
    add_borders(ws, ws.max_row, 6)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("KPIs")
    ws2.append(["Indicador", "Valor", "Meta", "Cumplimiento"])
    kpis = [
        ("Ventas Mensuales", 125000, 120000, 1.04),
        ("Margen Bruto", 0.42, 0.40, 1.05),
        ("Satisfaccion Cliente", 4.5, 4.2, 1.07),
        ("Rotacion Inventario", 6.2, 5.5, 1.13),
        ("Ticket Promedio", 348, 320, 1.09),
    ]
    for k in kpis:
        ws2.append(k)
    stylize_header(ws2, 4)
    add_borders(ws2, ws2.max_row, 4)

    ws3 = wb.create_sheet("Ventas por Producto")
    ws3.append(["Producto", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Total"])
    prods = ["Silla", "Escritorio", "Laptop", "Monitor", "Teclado", "Mouse", "Archivador"]
    for p in prods:
        row = [p]
        total = 0
        for _ in range(6):
            v = random.randint(5000, 30000)
            row.append(v)
            total += v
        row.append(total)
        ws3.append(row)
    stylize_header(ws3, 8)
    center_cols(ws3, ["B", "C", "D", "E", "F", "G", "H"])
    add_borders(ws3, ws3.max_row, 8)
    ws3.freeze_panes = "B2"

    wb.save(os.path.join(out, "datos_visualizacion.xlsx"))
    print(f"  OK: c07_visualizacion/datos_visualizacion.xlsx")

# ── c08_escenarios ────────────────────────────────────────────────────
def gen_c08():
    out = os.path.join(CODIGOS, "c08_escenarios")
    os.makedirs(out, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Negocio"

    # Input section
    ws["A1"] = "MODELO DE NEGOCIO"
    ws["A1"].font = Font(bold=True, size=14, color="0A7E5C")
    ws.merge_cells("A1:C1")

    vars_def = [
        ("Precio Unitario", 150, 120, 180),
        ("Cantidad Vendida", 5000, 3000, 8000),
        ("Costo Variable Unitario", 60, 50, 80),
        ("Costo Fijo Mensual", 200000, 150000, 250000),
    ]
    ws.append([])
    ws.append(["Variable", "Valor Base", "Optimista", "Pesimista"])
    for v, base, opt, pes in vars_def:
        ws.append([v, base, opt, pes])
    stylize_header(ws, 4)
    row_start = ws.max_row - len(vars_def)
    for r in range(row_start + 1, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)

    # Results section
    ws.append([])
    ws.append(["RESULTADOS"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12, color="0A7E5C")
    ws.append(["Concepto", "Base", "Optimista", "Pesimista"])
    stylize_header(ws, 4)
    for scenario in ["Base", "Optimista", "Pesimista"]:
        col = 2 if scenario == "Base" else 3 if scenario == "Optimista" else 4
        precio = 150 if scenario == "Base" else 180 if scenario == "Optimista" else 120
        cant = 5000 if scenario == "Base" else 8000 if scenario == "Optimista" else 3000
        cv = 60 if scenario == "Base" else 50 if scenario == "Optimista" else 80
        cf = 200000 if scenario == "Base" else 150000 if scenario == "Optimista" else 250000
        ing = precio * cant
        cv_total = cv * cant
        margen = ing - cv_total
        utilidad = margen - cf
        ws.cell(row=ws.max_row + 1, column=1).value = "Ingresos"
        ws.cell(row=ws.max_row, column=col).value = ing
        ws.cell(row=ws.max_row + 1, column=1).value = "Costo Variable Total"
        ws.cell(row=ws.max_row, column=col).value = cv_total
        ws.cell(row=ws.max_row + 1, column=1).value = "Margen Bruto"
        ws.cell(row=ws.max_row, column=col).value = margen
        ws.cell(row=ws.max_row + 1, column=1).value = "Costo Fijo"
        ws.cell(row=ws.max_row, column=col).value = cf
        ws.cell(row=ws.max_row + 1, column=1).value = "Utilidad Neta"
        ws.cell(row=ws.max_row, column=col).value = utilidad
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    add_borders(ws, ws.max_row, 4)
    wb.save(os.path.join(out, "modelo_negocio.xlsx"))
    print(f"  OK: c08_escenarios/modelo_negocio.xlsx")

# ── c09_powerquery ────────────────────────────────────────────────────
def gen_c09():
    out = os.path.join(CODIGOS, "c09_powerquery")
    os.makedirs(out, exist_ok=True)

    headers = ["Fecha", "ID_Producto", "Producto", "Cantidad", "PrecioUnitario", "Total", "Region"]
    productos_pq = [
        (101, "Silla", 350), (102, "Escritorio", 850), (103, "Laptop", 3200),
        (104, "Monitor", 1200), (105, "Teclado", 250), (106, "Mouse", 180),
    ]
    regiones = ["Lima", "Arequipa", "Cusco", "Trujillo"]

    for mes_idx, mes_nombre in enumerate(["ene", "feb", "mar"], 1):
        path = os.path.join(out, f"ventas_{mes_nombre}.csv")
        import csv
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for _ in range(200):
                pid, prod, pu = random.choice(productos_pq)
                cant = random.randint(1, 15)
                dia = random.randint(1, 28)
                fecha = f"2025-{mes_idx:02d}-{dia:02d}"
                w.writerow([fecha, pid, prod, cant, pu, round(cant * pu, 2), random.choice(regiones)])
        print(f"  OK: c09_powerquery/ventas_{mes_nombre}.csv")

    # productos.xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["ID_Producto", "Nombre", "Categoria", "Precio", "Proveedor"])
    prods = [
        (101, "Silla", "Muebles", 350, "Muebles SAC"),
        (102, "Escritorio", "Muebles", 850, "Muebles SAC"),
        (103, "Laptop", "Electrónicos", 3200, "TechGlobal"),
        (104, "Monitor", "Electrónicos", 1200, "TechGlobal"),
        (105, "Teclado", "Electrónicos", 250, "InOut Corp"),
        (106, "Mouse", "Electrónicos", 180, "InOut Corp"),
    ]
    for p in prods:
        ws.append(p)
    stylize_header(ws, 5)
    add_borders(ws, ws.max_row, 5)
    wb.save(os.path.join(out, "productos.xlsx"))
    print(f"  OK: c09_powerquery/productos.xlsx")

# ── c10_proyecto ──────────────────────────────────────────────────────
def gen_c10():
    out = os.path.join(CODIGOS, "c10_proyecto")
    os.makedirs(out, exist_ok=True)

    # ventas_mes.csv — current month
    headers = ["Fecha", "ID_Producto", "ID_Cliente", "Cantidad", "PrecioUnitario", "Total"]
    with open(os.path.join(out, "ventas_mes.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for i in range(1, 501):
            pid = random.randint(101, 110)
            cli = random.randint(201, 210)
            cant = random.randint(1, 20)
            pu = random.uniform(25, 3200)
            w.writerow([f"2025-06-{random.randint(1,30):02d}", pid, cli, cant, round(pu, 2), round(cant * pu, 2)])
    print(f"  OK: c10_proyecto/ventas_mes.csv")

    # ventas_historico.xlsx — 12 months of previous data
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Fecha", "ID_Producto", "Producto", "Categoria", "ID_Cliente", "Cantidad", "Monto", "Region"])
    prods_hist = [
        (101, "Silla", "Muebles"), (102, "Escritorio", "Muebles"), (103, "Laptop", "Electrónicos"),
        (104, "Monitor", "Electrónicos"), (105, "Teclado", "Electrónicos"), (106, "Mouse", "Electrónicos"),
        (107, "Archivador", "Oficina"), (108, "Papel Bond", "Oficina"),
    ]
    regs = ["Lima", "Arequipa", "Cusco", "Trujillo", "Piura"]
    for mes in range(1, 13):
        for _ in range(200):
            pid, prod, cat = random.choice(prods_hist)
            cant = random.randint(1, 30)
            pu = random.uniform(25, 3200)
            dia = random.randint(1, 28)
            ws.append([
                f"2025-{mes:02d}-{dia:02d}", pid, prod, cat,
                random.randint(201, 210), cant, round(cant * pu, 2), random.choice(regs),
            ])
    stylize_header(ws, 8)
    add_borders(ws, ws.max_row, 8)
    ws.freeze_panes = "A2"
    wb.save(os.path.join(out, "ventas_historico.xlsx"))
    print(f"  OK: c10_proyecto/ventas_historico.xlsx")

    # productos.xlsx
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Productos"
    ws2.append(["ID_Producto", "Nombre", "Categoria", "Precio", "Costo", "Proveedor"])
    for pid, prod, cat in prods_hist:
        precio = {"Silla": 350, "Escritorio": 850, "Laptop": 3200, "Monitor": 1200,
                  "Teclado": 250, "Mouse": 180, "Archivador": 65, "Papel Bond": 25}[prod]
        costo = int(precio * random.uniform(0.4, 0.7))
        prov = random.choice(["Proveedor A", "Proveedor B", "Proveedor C"])
        ws2.append([pid, prod, cat, precio, costo, prov])
    stylize_header(ws2, 6)
    add_borders(ws2, ws2.max_row, 6)
    wb2.save(os.path.join(out, "productos.xlsx"))
    print(f"  OK: c10_proyecto/productos.xlsx")

    # clientes.xlsx
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Clientes"
    ws3.append(["ID_Cliente", "Nombre", "Segmento", "Region", "AntiguedadMeses"])
    segmentos = ["VIP", "Premium", "Regular", "Básico"]
    for cid in range(201, 211):
        ws3.append([cid, f"Cliente {cid}", random.choice(segmentos),
                    random.choice(regs), random.randint(1, 60)])
    stylize_header(ws3, 5)
    add_borders(ws3, ws3.max_row, 5)
    wb3.save(os.path.join(out, "clientes.xlsx"))
    print(f"  OK: c10_proyecto/clientes.xlsx")


def main():
    os.makedirs(CODIGOS, exist_ok=True)
    print("Generando archivos de ejemplo...")
    gen_c01()
    gen_c02()
    gen_c03()
    gen_c04()
    gen_c05()
    gen_c06()
    gen_c07()
    gen_c08()
    gen_c09()
    gen_c10()
    print("\nTodos los archivos generados exitosamente.")

if __name__ == "__main__":
    main()
