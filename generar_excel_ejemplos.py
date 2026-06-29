#!/usr/bin/env python3
"""
Genera archivos .xlsx de ejemplo en las carpetas codigos/ de
excel_basico, excel_intermedio y excel_avanzado, uno por capitulo.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, IconSetRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from copy import copy

BASE = Path(r"C:\Users\alexg\OneDrive\Alex_2026\libros epub")

# ── helpers ──────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
S_FILL = PatternFill("solid", fgColor="FFF2CC")


def style_header(ws, row=1, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def add_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


# ═══════════════════════════════════════════════════════════════════════
#  EXCEL BÁSICO
# ═══════════════════════════════════════════════════════════════════════

def basico_cap1(wb):
    ws = wb.active
    ws.title = "Interfaz"
    ws["A1"] = "Producto"; ws["B1"] = "Cantidad"; ws["C1"] = "Unidad"
    ws["D1"] = "Precio Unit."; ws["E1"] = "Proveedor"
    data = [
        ["Madera Cedro", 50, "tablas", 45, "Maderas Pérez"],
        ["Madera Roble", 30, "tablas", 62, "Maderas Pérez"],
        ["Madera Pino", 100, "tablas", 22, "El Maderero"],
        ["Triplay", 40, "planchas", 35, "Maderas Pérez"],
        ["Barniz mate", 10, "galones", 28, "Pinturas Unidas"],
        ["Tornillos #10", 500, "unidades", 0.5, "Ferretería El Tornillo"],
        ["Lija fina", 200, "hojas", 1.2, "Ferretería El Tornillo"],
        ["Pegamento madera", 12, "botellas", 15, "Pinturas Unidas"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 5)
    add_border_range(ws, 1, len(data) + 1, 1, 5)
    ws["A12"] = "Nota: Identifica las partes de Excel: cinta de opciones, barra de fórmulas, celdas, hojas."
    ws["A12"].font = Font(italic=True, color="666666")
    auto_width(ws)


def basico_cap2(wb):
    ws = wb.active
    ws.title = "Tipos de Datos"
    ws["A1"] = "Producto"; ws["B1"] = "Cantidad"; ws["C1"] = "Precio"
    ws["D1"] = "Fecha Ingreso"; ws["E1"] = "Calidad"
    data = [
        ["Cinceles", 15, 25, "15/03/2026", "A"],
        ["Martillos", 8, 45, "15/03/2026", "A"],
        ["Cepillos carpintero", 6, 60, "18/03/2026", "B"],
        ["Escuadras", 12, 18, "20/03/2026", "B"],
        ["Niveles", 4, 35, "20/03/2026", "C"],
        ["Taladro inalámbrico", 2, 280, "22/03/2026", "A"],
        ["Sierra circular", 3, 350, "22/03/2026", "A"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 5)
    add_border_range(ws, 1, len(data) + 1, 1, 5)
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=4).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=3).number_format = '#,##0.00'
    auto_width(ws)


def basico_cap3(wb):
    ws = wb.active
    ws.title = "Fórmulas Básicas"
    ws["A1"] = "Producto"; ws["B1"] = "Madera"; ws["C1"] = "Barniz"
    ws["D1"] = "Tornillos"; ws["E1"] = "Mano Obra"; ws["F1"] = "Costo Total"
    ws["G1"] = "Precio Venta"; ws["H1"] = "Ganancia"; ws["I1"] = "% Ganancia"
    data = [
        ["Silla cedro", 45, 7, 3, 35, None, 150, None, None],
        ["Mesa roble", 186, 14, 8, 70, None, 350, None, None],
        ["Banquito pino", 22, 5, 2, 20, None, 80, None, None],
        ["Estante triplay", 70, 10, 5, 45, None, 180, None, None],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=6).value = f"=B{i}+C{i}+D{i}+E{i}"
        ws.cell(row=i, column=8).value = f"=G{i}-F{i}"
        ws.cell(row=i, column=9).value = f"=(G{i}-F{i})/F{i}"
        ws.cell(row=i, column=9).number_format = "0.00%"
    style_header(ws, 1, 9)
    add_border_range(ws, 1, 5, 1, 9)
    for r in range(2, 6):
        for col in (6, 7, 8):
            ws.cell(row=r, column=col).number_format = '#,##0.00'
    auto_width(ws)


def basico_cap4(wb):
    ws = wb.active
    ws.title = "Referencias"
    ws["E1"] = "Tipo Cambio"; ws["F1"] = 3.75
    ws["A3"] = "Producto"; ws["B3"] = "Precio S/"; ws["C3"] = "IVA (18%)"
    ws["D3"] = "Precio Final"; ws["E3"] = "Precio USD"
    data = [
        ["Silla cedro", 150],
        ["Mesa roble", 350],
        ["Banquito pino", 80],
        ["Estante triplay", 180],
    ]
    for i, row in enumerate(data, 4):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
        ws.cell(row=i, column=3).value = f"=B{i}*$F$1"
        ws.cell(row=i, column=4).value = f"=B{i}+C{i}"
        ws.cell(row=i, column=5).value = f"=B{i}*$E$1"
    style_header(ws, 3, 5)
    add_border_range(ws, 3, 7, 1, 5)
    for r in range(4, 8):
        for col in range(2, 6):
            ws.cell(row=r, column=col).number_format = '#,##0.00'
    ws["E1"].number_format = '#,##0.00'
    ws["A9"] = "Nota: $F$1 = IVA (absoluta), $E$1 = tipo de cambio (absoluta)"
    ws["A9"].font = Font(italic=True, color="666666")
    auto_width(ws)


def basico_cap5(wb):
    ws = wb.active
    ws.title = "Formato Celdas"
    ws["A1"] = "Producto"; ws["B1"] = "Cantidad"; ws["C1"] = "Precio Unit."
    ws["D1"] = "Fecha Ingreso"; ws["E1"] = "Calidad"
    data = [
        ["Madera Cedro", 50, 45, "15/01/2026", "A"],
        ["Madera Roble", 30, 62, "15/01/2026", "A"],
        ["Madera Pino", 100, 22, "20/01/2026", "B"],
        ["Triplay", 40, 35, "20/01/2026", "C"],
        ["Barniz mate", 10, 28, "10/02/2026", "C"],
        ["Tornillos #10", 500, 0.5, "05/02/2026", "B"],
        ["Lija fina", 200, 1.2, "05/02/2026", "A"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 5)
    add_border_range(ws, 1, len(data) + 1, 1, 5)
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=3).number_format = 'S/#,##0.00'
        ws.cell(row=r, column=4).number_format = '[$-es-PE]dd "de" mmmm "de" aaaa'
    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
    RED_FILL = PatternFill("solid", fgColor="FFC7CE")
    b_range = f"B2:B{len(data)+1}"
    ws.conditional_formatting.add(b_range,
        DataBarRule(start_type="min", end_type="max", color="5B9BD5"))
    ws.conditional_formatting.add(b_range,
        ColorScaleRule(start_type="min", start_color="FFC7CE",
                       mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                       end_type="max", end_color="C6EFCE"))
    auto_width(ws)


def basico_cap6(wb):
    ws = wb.active
    ws.title = "Funciones"
    headers = ["Mes", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Total"]
    ws.append(headers)
    data = [
        ["Silla", 1200, 1350, 1100, 1400, 1300, 1450],
        ["Mesa", 800, 750, 900, 850, 780, 920],
        ["Estante", 450, 500, 480, 520, 490, 510],
        ["Ropero", 600, 580, 620, 640, 590, 610],
    ]
    for row in data:
        ws.append(row + [None])
    for r in range(2, 6):
        ws.cell(row=r, column=8).value = f"=SUM(B{r}:G{r})"
        ws.cell(row=r, column=8).number_format = '#,##0'
    style_header(ws, 1, 8)
    add_border_range(ws, 1, 5, 1, 8)
    r = 7
    ws.cell(row=r, column=1, value="PROMEDIO").font = Font(bold=True)
    for c in range(2, 9):
        ws.cell(row=r, column=c).value = f"=AVERAGE({get_column_letter(c)}2:{get_column_letter(c)}5)"
    r = 8
    ws.cell(row=r, column=1, value="MAX").font = Font(bold=True)
    for c in range(2, 9):
        ws.cell(row=r, column=c).value = f"=MAX({get_column_letter(c)}2:{get_column_letter(c)}5)"
    r = 9
    ws.cell(row=r, column=1, value="MIN").font = Font(bold=True)
    for c in range(2, 9):
        ws.cell(row=r, column=c).value = f"=MIN({get_column_letter(c)}2:{get_column_letter(c)}5)"
    r = 10
    ws.cell(row=r, column=1, value="CONTAR").font = Font(bold=True)
    for c in range(2, 9):
        ws.cell(row=r, column=c).value = f"=COUNT({get_column_letter(c)}2:{get_column_letter(c)}5)"
    add_border_range(ws, 7, 10, 1, 8)
    auto_width(ws)


def basico_cap7(wb):
    ws = wb.active
    ws.title = "Gráficos"
    ws["A1"] = "Producto"; ws["B1"] = "Total Ventas"
    data = [["Silla", 7800], ["Mesa", 5000], ["Estante", 2950], ["Ropero", 3640]]
    for i, row in enumerate(data, 2):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
    style_header(ws, 1, 2)
    add_border_range(ws, 1, 5, 1, 2)
    chart = BarChart()
    chart.title = "Ventas por Producto"
    chart.style = 10
    chart.y_axis.title = "Soles"
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "D1")
    ws["A8"] = "Meses"
    for i, m in enumerate(["Ene", "Feb", "Mar", "Abr", "May", "Jun"], 1):
        ws.cell(row=8, column=i + 1, value=m)
    ws["A9"] = "Totales"
    for i, v in enumerate([3050, 3180, 3100, 3410, 3160, 3490], 1):
        ws.cell(row=9, column=i + 1, value=v)
    chart2 = LineChart()
    chart2.title = "Tendencia Mensual"
    data_ref2 = Reference(ws, min_col=2, min_row=8, max_row=9)
    cats2 = Reference(ws, min_col=2, max_col=7, min_row=8)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "D18")
    auto_width(ws)


def basico_cap8(wb):
    ws = wb.active
    ws.title = "Ordenar y Filtrar"
    ws["A1"] = "Código"; ws["B1"] = "Producto"; ws["C1"] = "Precio"
    ws["D1"] = "Calidad"; ws["E1"] = "Categoría"
    data = [
        ["CAT-001", "Banquito", 80, "B", "Accesorios"],
        ["CAT-002", "Cama king", 680, "A", "Dormitorio"],
        ["CAT-003", "Cama queen", 450, "A", "Dormitorio"],
        ["CAT-004", "Cómoda", 420, "A", "Dormitorio"],
        ["CAT-005", "Escritorio", 280, "B", "Sala"],
        ["CAT-006", "Estante", 180, "C", "Sala"],
        ["CAT-007", "Mesa centro", 250, "B", "Sala"],
        ["CAT-008", "Mesa noche", 120, "B", "Dormitorio"],
        ["CAT-009", "Mesa comedor", 520, "A", "Comedor"],
        ["CAT-010", "Ropero", 890, "A", "Dormitorio"],
        ["CAT-011", "Silla", 150, "B", "Comedor"],
        ["CAT-012", "Sillón", 520, "A", "Sala"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 5)
    add_border_range(ws, 1, len(data) + 1, 1, 5)
    ws.auto_filter.ref = f"A1:E{len(data)+1}"
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=3).number_format = '#,##0'
    auto_width(ws)


def basico_cap9(wb):
    ws = wb.active
    ws.title = "Impresión"
    ws["A1"] = "Código"; ws["B1"] = "Producto"; ws["C1"] = "Cant."
    ws["D1"] = "P. Unit."; ws["E1"] = "Subtotal"; ws["F1"] = "Dto."
    ws["G1"] = "Total"
    data = [
        ["DORM-01", "Cama king cedro", 2, 680],
        ["DORM-02", "Mesa de noche", 4, 120],
        ["DORM-03", "Cómoda", 2, 420],
        ["SALA-01", "Sillón", 3, 520],
        ["SALA-02", "Mesa de centro", 1, 250],
        ["COMED-01", "Mesa comedor", 1, 520],
        ["COMED-02", "Silla comedor", 8, 150],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=5).value = f"=C{i}*D{i}"
        ws.cell(row=i, column=6).value = 0.1 if i <= 5 else 0
        ws.cell(row=i, column=6).number_format = "0%"
        ws.cell(row=i, column=7).value = f"=E{i}*(1-F{i})"
        ws.cell(row=i, column=7).number_format = '#,##0.00'
    ws.cell(row=9, column=1, value="TOTAL GENERAL").font = Font(bold=True, size=12)
    ws.cell(row=9, column=7).value = "=SUM(G2:G8)"
    ws.cell(row=9, column=7).number_format = '#,##0.00'
    style_header(ws, 1, 7)
    add_border_range(ws, 1, 9, 1, 7)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 1.5
    ws.page_margins.right = 1.5
    ws.page_margins.top = 1.5
    ws.page_margins.bottom = 1.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    auto_width(ws)


def basico_cap10(wb):
    ws = wb.active
    ws.title = "Proyecto Final"
    ws["A1"] = "Indicador"; ws["B1"] = "Valor"; ws["C1"] = "Meta"; ws["D1"] = "Estado"
    data = [
        ["Ventas del mes", 28450, 25000, "Meta superada"],
        ["Costos totales", 18200, 17500, "Por ajustar"],
        ["Ganancia neta", 10250, 7500, "Excelente"],
        ["Pedidos pendientes", 8, 5, "Mayor capacidad"],
        ["Días entrega prom.", 12, 10, "Mejorando"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 4)
    add_border_range(ws, 1, 6, 1, 4)
    for r in range(2, 7):
        ws.cell(row=r, column=2).number_format = '#,##0'
        ws.cell(row=r, column=3).number_format = '#,##0'
    ws["A8"] = "Producto"; ws["B8"] = "Cant."; ws["C8"] = "P. Unit."
    ws["D8"] = "Total"
    inv = [["Cedro", 50, 45], ["Roble", 30, 62], ["Pino", 100, 22],
           ["Triplay", 40, 35], ["Barniz", 10, 28]]
    for i, row in enumerate(inv, 9):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
        ws.cell(row=i, column=3, value=row[2])
        ws.cell(row=i, column=4).value = f"=B{i}*C{i}"
        ws.cell(row=i, column=4).number_format = '#,##0.00'
    style_header(ws, 8, 4)
    add_border_range(ws, 8, 13, 1, 4)
    auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════
#  EXCEL INTERMEDIO
# ═══════════════════════════════════════════════════════════════════════

def intermedio_cap1(wb):
    ws = wb.active
    ws.title = "Formato Condicional"
    ws["A1"] = "Producto"; ws["B1"] = "Ventas Ene"; ws["C1"] = "Ventas Feb"
    ws["D1"] = "Ventas Mar"; ws["E1"] = "Total"
    data = [
        ["Silla cedro", 1200, 1350, 1100, None],
        ["Mesa roble", 800, 750, 900, None],
        ["Banquito pino", 450, 500, 480, None],
        ["Estante triplay", 600, 580, 620, None],
        ["Ropero", 900, 850, 920, None],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=5).value = f"=SUM(B{i}:D{i})"
    style_header(ws, 1, 5)
    add_border_range(ws, 1, 6, 1, 5)
    ws.conditional_formatting.add(f"B2:D6",
        DataBarRule(start_type="min", end_type="max", color="5B9BD5"))
    ws.conditional_formatting.add(f"B2:D6",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))
    ws.conditional_formatting.add(f"B2:D6",
        IconSetRule(icon_style="3TrafficLights1", type="percent", values=[0, 33, 67], showValue=True))
    auto_width(ws)


def intermedio_cap2(wb):
    ws = wb.active
    ws.title = "Validación Datos"
    ws["A1"] = "Fecha"; ws["B1"] = "Producto"; ws["C1"] = "Cantidad"
    ws["D1"] = "Cliente"; ws["E1"] = "Tipo Pago"
    data = [
        ["01/06/2026", "Silla cedro", 2, "Sr. García", "Contado"],
        ["02/06/2026", "Mesa roble", 1, "Sra. López", "Crédito"],
        ["03/06/2026", "Estante triplay", 5, "Hotel Miraflores", "Contado"],
        ["04/06/2026", "Cama queen", 1, "Sra. María", "Crédito"],
        ["05/06/2026", "Silla cedro", 4, "Rest. El Buen Sabor", "Contado"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 5)
    add_border_range(ws, 1, 6, 1, 5)
    dv = DataValidation(type="list", formula1='"Contado,Crédito,Transferencia"', allow_blank=True)
    dv.error = "Selecciona un tipo de pago válido"
    dv.errorTitle = "Pago inválido"
    dv.prompt = "Selecciona el tipo de pago"
    dv.promptTitle = "Tipo de pago"
    ws.add_data_validation(dv)
    dv.add("E2:E6")
    auto_width(ws)


def intermedio_cap3(wb):
    ws = wb.active
    ws.title = "Búsquedas"
    ws["A1"] = "ID"; ws["B1"] = "Producto"; ws["C1"] = "Precio"
    ws["D1"] = "Categoría"
    prods = [
        [101, "Silla cedro", 150, "Comedor"],
        [102, "Mesa roble", 350, "Comedor"],
        [103, "Banquito pino", 80, "Accesorios"],
        [104, "Estante triplay", 180, "Sala"],
        [105, "Cama queen", 450, "Dormitorio"],
        [106, "Ropero", 890, "Dormitorio"],
    ]
    for i, row in enumerate(prods, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 4)
    add_border_range(ws, 1, len(prods) + 1, 1, 4)
    ws["A9"] = "ID a buscar"; ws["B9"] = 103
    ws["A10"] = "Producto"; ws["B10"] = '=VLOOKUP(B9,A2:D7,2,FALSE)'
    ws["A11"] = "Precio"; ws["B11"] = '=VLOOKUP(B9,A2:D7,3,FALSE)'
    for r in range(2, len(prods) + 2):
        ws.cell(row=r, column=3).number_format = '#,##0.00'
    auto_width(ws)


def intermedio_cap4(wb):
    ws = wb.active
    ws.title = "Tablas Dinámicas"
    ws["A1"] = "Fecha"; ws["B1"] = "Producto"; ws["C1"] = "Categoría"
    ws["D1"] = "Cantidad"; ws["E1"] = "Venta"
    import random
    random.seed(42)
    cats = {"Silla cedro": "Comedor", "Mesa roble": "Comedor",
            "Banquito": "Accesorios", "Estante": "Sala",
            "Cama queen": "Dormitorio", "Ropero": "Dormitorio"}
    prods = list(cats.keys())
    data = []
    for dia in range(1, 31):
        p = random.choice(prods)
        data.append([f"01/{dia:02d}/2026", p, cats[p],
                     random.randint(1, 5), random.randint(50, 900)])
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 5)
    add_border_range(ws, 1, 31, 1, 5)
    auto_width(ws)


def intermedio_cap5(wb):
    ws = wb.active
    ws.title = "Gráficos Avanzados"
    ws["A1"] = "Mes"; ws["B1"] = "Ventas (S/)"; ws["C1"] = "Gastos (S/)"
    ws["D1"] = "% Crecimiento"
    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun"]
    ventas = [12000, 13500, 11000, 14000, 13000, 14500]
    gastos = [8000, 8500, 7800, 9000, 8200, 8800]
    for i, m in enumerate(meses, 2):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=ventas[i - 2])
        ws.cell(row=i, column=3, value=gastos[i - 2])
        if i > 2:
            ws.cell(row=i, column=4).value = f"=(B{i}-B{i-1})/B{i-1}"
            ws.cell(row=i, column=4).number_format = "0.00%"
    style_header(ws, 1, 4)
    add_border_range(ws, 1, 7, 1, 4)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Ventas vs Gastos"
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=7)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A9")
    for r in range(2, 8):
        ws.cell(row=r, column=2).number_format = '#,##0'
        ws.cell(row=r, column=3).number_format = '#,##0'
    auto_width(ws)


def intermedio_cap6(wb):
    ws = wb.active
    ws.title = "Limpieza Datos"
    ws["A1"] = "Nombre Completo"; ws["B1"] = "Email"; ws["C1"] = "Teléfono"
    ws["D1"] = "Ciudad"
    data = [
        ["Sofía Mendoza", "sofia.mendoza@email.com", "999-888-777", "Lima"],
        ["Carlos Pérez", "carlos.perez@email.com", "999-888-776", "Lima"],
        ["SOFÍA MENDOZA", "sofia.mendoza@email.com", "999-888-777", "LIMA"],
        ["María García", "maria.garcia@email.com", "999-888-775", "Arequipa"],
        ["Pedro Huamán", "pedro.h@email.com", "999-888-774", "Cusco"],
        ["sofía mendoza", "sofia.mendoza@email.com", "999-888-777", "lima"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 4)
    add_border_range(ws, 1, len(data) + 1, 1, 4)
    ws["A9"] = "Nota: Contiene duplicados y datos inconsistentes. Usa Quitar Duplicados y Texto en Columnas."
    ws["A9"].font = Font(italic=True, color="666666")
    auto_width(ws)


def intermedio_cap7(wb):
    ws = wb.active
    ws.title = "Rangos Nombrados"
    ws["A1"] = "Producto"; ws["B1"] = "Ene"; ws["C1"] = "Feb"; ws["D1"] = "Mar"
    data = [
        ["Silla", 1200, 1350, 1100],
        ["Mesa", 800, 750, 900],
        ["Estante", 450, 500, 480],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 4)
    add_border_range(ws, 1, 4, 1, 4)
    from openpyxl.workbook.defined_name import DefinedName
    dn = DefinedName("VentasEne", attr_text="RangosNombrados!$B$2:$B$4")
    wb.defined_names.add(dn)
    ws["A6"] = "Producto:"; ws["B6"] = "Silla"
    ws["A7"] = "Venta Ene:"; ws["B7"] = "=VentasEne"
    auto_width(ws)


def intermedio_cap8(wb):
    ws = wb.active
    ws.title = "Análisis Hipotético"
    ws["A1"] = "Escenario"; ws["B1"] = "Precio Unit."; ws["C1"] = "Cantidad"
    ws["D1"] = "Costo Fijo"; ws["E1"] = "Ingreso"; ws["F1"] = "Costo Total"
    ws["G1"] = "Utilidad"
    escenarios = [
        ["Optimista", 250, 100, 5000, None, None, None],
        ["Realista", 200, 80, 5000, None, None, None],
        ["Pesimista", 150, 60, 5000, None, None, None],
    ]
    for i, row in enumerate(escenarios, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=5).value = f"=B{i}*C{i}"
        ws.cell(row=i, column=6).value = f"=C{i}*100+{5000}"
        ws.cell(row=i, column=7).value = f"=E{i}-F{i}"
    style_header(ws, 1, 7)
    add_border_range(ws, 1, 4, 1, 7)
    auto_width(ws)


def intermedio_cap9(wb):
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "Indicador"; ws["B1"] = "Valor"
    kpis = [
        ["Ventas Totales", 28450],
        ["Costo Total", 18200],
        ["Ganancia Neta", 10250],
        ["Pedidos", 8],
        ["Clientes", 12],
    ]
    for i, row in enumerate(kpis, 2):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
    style_header(ws, 1, 2)
    add_border_range(ws, 1, 6, 1, 2)
    for r in range(2, 7):
        ws.cell(row=r, column=2).number_format = '#,##0'
    ws["A8"] = "Mes"; ws["B8"] = "Ventas"
    ventas = [["Ene", 12000], ["Feb", 13500], ["Mar", 11000],
              ["Abr", 14000], ["May", 13000], ["Jun", 14500]]
    for i, row in enumerate(ventas, 9):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
    chart = BarChart()
    chart.title = "Tendencia Ventas"
    data_ref = Reference(ws, min_col=2, min_row=8, max_row=14)
    cats = Reference(ws, min_col=1, min_row=9, max_row=14)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D1")
    auto_width(ws)


def intermedio_cap10(wb):
    ws = wb.active
    ws.title = "Proyecto Final"
    ws["A1"] = "Fecha"; ws["B1"] = "Transacción"; ws["C1"] = "Tipo"
    ws["D1"] = "Monto"; ws["E1"] = "Saldo"
    trans = [
        ["01/01/2026", "Saldo inicial", "Apertura", 50000, None],
        ["05/01/2026", "Venta sillas", "Ingreso", 12000, None],
        ["10/01/2026", "Compra madera", "Egreso", -8000, None],
        ["15/01/2026", "Venta mesas", "Ingreso", 15000, None],
        ["20/01/2026", "Pago alquiler", "Egreso", -2500, None],
        ["25/01/2026", "Venta estantes", "Ingreso", 9000, None],
    ]
    for i, row in enumerate(trans, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
        if i == 2:
            ws.cell(row=i, column=5).value = f"=D{i}"
        else:
            ws.cell(row=i, column=5).value = f"=E{i-1}+D{i}"
        ws.cell(row=i, column=5).number_format = '#,##0.00'
    style_header(ws, 1, 5)
    add_border_range(ws, 1, len(trans) + 1, 1, 5)
    for r in range(2, len(trans) + 2):
        ws.cell(row=r, column=4).number_format = '#,##0.00'
    auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════
#  EXCEL AVANZADO
# ═══════════════════════════════════════════════════════════════════════

def avanzado_cap1(wb):
    ws = wb.active
    ws.title = "Power Query"
    ws["A1"] = "Fecha"; ws["B1"] = "Cuenta"; ws["C1"] = "Descripción"
    ws["D1"] = "Monto"; ws["E1"] = "Tipo"
    data = [
        ["01/01/2026", "Cta. 1001", "Depósito nómina", 5000, "Ingreso"],
        ["03/01/2026", "Cta. 1001", "Retiro ATM", -500, "Egreso"],
        ["05/01/2026", "Cta. 1002", "Transferencia recibida", 3000, "Ingreso"],
        ["07/01/2026", "Cta. 1001", "Pago electricidad", -250, "Egreso"],
        ["10/01/2026", "Cta. 1002", "Depósito efectivo", 2000, "Ingreso"],
        ["12/01/2026", "Cta. 1001", "Compra supermercado", -180, "Egreso"],
        ["15/01/2026", "Cta. 1003", "Transferencia saliente", -1000, "Egreso"],
        ["20/01/2026", "Cta. 1001", "Depósito cliente", 4500, "Ingreso"],
        ["25/01/2026", "Cta. 1002", "Pago servicio", -320, "Egreso"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 5)
    add_border_range(ws, 1, len(data) + 1, 1, 5)
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=4).number_format = '#,##0.00'
    ws["A12"] = "Nota: Datos listos para transformar con Power Query (Obtener y Transformar)"
    ws["A12"].font = Font(italic=True, color="666666")
    auto_width(ws)


def avanzado_cap2(wb):
    ws = wb.active
    ws.title = "Power Pivot"
    ws["A1"] = "Orden ID"; ws["B1"] = "Producto"; ws["C1"] = "Cantidad"
    ws["D1"] = "Precio Unit."; ws["E1"] = "Cliente"; ws["F1"] = "Fecha"
    data = [
        [1001, "Silla cedro", 10, 150, "Hotel Miraflores", "01/06/2026"],
        [1002, "Mesa roble", 5, 350, "Hotel Miraflores", "01/06/2026"],
        [1003, "Cama queen", 8, 450, "Clínica San Pablo", "05/06/2026"],
        [1004, "Estante triplay", 12, 180, "Hotel Miraflores", "05/06/2026"],
        [1005, "Ropero", 6, 890, "Clínica San Pablo", "10/06/2026"],
        [1006, "Silla cedro", 20, 150, "Municipalidad Lima", "15/06/2026"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 6)
    add_border_range(ws, 1, len(data) + 1, 1, 6)
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=4).number_format = '#,##0.00'
    auto_width(ws)


def avanzado_cap3(wb):
    ws = wb.active
    ws.title = "DAX Básico"
    ws["A1"] = "Producto"; ws["B1"] = "Cantidad Vendida"; ws["C1"] = "Precio Prom."
    ws["D1"] = "Ingreso Total"
    data = [
        ["Silla cedro", 120, 150, None],
        ["Mesa roble", 45, 350, None],
        ["Cama queen", 30, 450, None],
        ["Estante triplay", 80, 180, None],
        ["Ropero", 25, 890, None],
        ["Banquito pino", 60, 80, None],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=4).value = f"=B{i}*C{i}"
    style_header(ws, 1, 4)
    add_border_range(ws, 1, len(data) + 1, 1, 4)
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=4).number_format = '#,##0.00'
    ws["A9"] = "Total Ingresos"; ws["B9"] = "=SUM(D2:D7)"
    ws["B9"].number_format = '#,##0.00'
    ws["A10"] = "Promedio x Producto"; ws["B10"] = "=AVERAGE(D2:D7)"
    ws["B10"].number_format = '#,##0.00'
    ws["A11"] = "Producto Estrella"; ws["B11"] = "=MAX(D2:D7)"
    ws["B11"].number_format = '#,##0.00'
    auto_width(ws)


def avanzado_cap4(wb):
    ws = wb.active
    ws.title = "VBA y Macros"
    ws["A1"] = "Producto"; ws["B1"] = "Precio"; ws["C1"] = "Descuento%"
    ws["D1"] = "Precio Final"
    data = [
        ["Silla cedro", 150, 10, None],
        ["Mesa roble", 350, 15, None],
        ["Cama queen", 450, 5, None],
        ["Estante triplay", 180, 20, None],
        ["Ropero", 890, 10, None],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=4).value = f"=B{i}*(1-C{i}/100)"
    style_header(ws, 1, 4)
    add_border_range(ws, 1, len(data) + 1, 1, 4)
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=2).number_format = '#,##0.00'
        ws.cell(row=r, column=3).number_format = '0'
        ws.cell(row=r, column=4).number_format = '#,##0.00'
    ws["A8"] = "Macro sugerida: AplicarDescuento() - selecciona rango y aplica 10%"
    ws["A8"].font = Font(italic=True, color="666666")
    auto_width(ws)


def avanzado_cap5(wb):
    ws = wb.active
    ws.title = "Matrices Dinámicas"
    ws["A1"] = "Producto"; ws["B1"] = "Ene"; ws["C1"] = "Feb"; ws["D1"] = "Mar"
    ws["E1"] = "Abr"; ws["F1"] = "May"; ws["G1"] = "Jun"
    data = [
        ["Silla cedro", 1200, 1350, 1100, 1400, 1300, 1450],
        ["Mesa roble", 800, 750, 900, 850, 780, 920],
        ["Estante triplay", 450, 500, 480, 520, 490, 510],
        ["Ropero", 600, 580, 620, 640, 590, 610],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 7)
    add_border_range(ws, 1, len(data) + 1, 1, 7)
    ws["A7"] = "Totales"
    for c in range(2, 8):
        ws.cell(row=7, column=c).value = f"=SUM({get_column_letter(c)}2:{get_column_letter(c)}5)"
    ws["A8"] = "Promedio"
    for c in range(2, 8):
        ws.cell(row=8, column=c).value = f"=AVERAGE({get_column_letter(c)}2:{get_column_letter(c)}5)"
    ws["A9"] = "Nota: Usa FÓRMULAS DINÁMICAS con FILTER, UNIQUE, SORT para análisis avanzado"
    ws["A9"].font = Font(italic=True, color="666666")
    auto_width(ws)


def avanzado_cap6(wb):
    ws = wb.active
    ws.title = "Solver"
    ws["A1"] = "Variable"; ws["B1"] = "Valor"
    ws["A2"] = "Prod. Sillas"; ws["B2"] = 50
    ws["A3"] = "Prod. Mesas"; ws["B3"] = 30
    ws["A4"] = "Prod. Estantes"; ws["B4"] = 20
    ws["A6"] = "Restricciones"
    ws["A7"] = "Madera usada (tablas)"; ws["B7"] = "=B2*2+B3*4+B4*3"
    ws["A8"] = "Horas MO"; ws["B8"] = "=B2*3+B3*5+B4*2"
    ws["A9"] = "Madera disponible"; ws["C9"] = 300
    ws["A10"] = "Horas disponibles"; ws["C10"] = 400
    ws["A12"] = "Ganancia Unitaria"
    ws["B12"] = 60; ws["C12"] = 120; ws["D12"] = 45
    ws["A13"] = "Ganancia Total"; ws["B13"] = "=B2*B12+B3*C12+B4*D12"
    style_header(ws, 1, 4)
    auto_width(ws)


def avanzado_cap7(wb):
    ws = wb.active
    ws.title = "Dashboard Avanzado"
    ws["A1"] = "KPI"; ws["B1"] = "Valor"; ws["C1"] = "Meta"
    ws["D1"] = "Variación"
    kpis = [
        ["Ingresos", 284500, 250000, None],
        ["Costos", 182000, 175000, None],
        ["Margen %", 36, 30, None],
        ["Clientes", 48, 40, None],
    ]
    for i, row in enumerate(kpis, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=4).value = f"=(B{i}-C{i})/C{i}"
        ws.cell(row=i, column=4).number_format = "0.00%"
    style_header(ws, 1, 4)
    add_border_range(ws, 1, 5, 1, 4)
    for r in range(2, 6):
        ws.cell(row=r, column=2).number_format = '#,##0'
        ws.cell(row=r, column=3).number_format = '#,##0'
    auto_width(ws)


def avanzado_cap8(wb):
    ws = wb.active
    ws.title = "Power BI"
    ws["A1"] = "Año"; ws["B1"] = "Trimestre"; ws["C1"] = "Producto"
    ws["D1"] = "Canal"; ws["E1"] = "Ventas"
    data = [
        [2025, "Q1", "Silla cedro", "Online", 35000],
        [2025, "Q1", "Mesa roble", "Online", 28000],
        [2025, "Q2", "Silla cedro", "Tienda", 42000],
        [2025, "Q2", "Mesa roble", "Online", 31000],
        [2025, "Q3", "Silla cedro", "Tienda", 38000],
        [2025, "Q3", "Mesa roble", "Tienda", 29000],
        [2025, "Q4", "Silla cedro", "Online", 45000],
        [2025, "Q4", "Mesa roble", "Tienda", 33000],
        [2026, "Q1", "Silla cedro", "Online", 48000],
        [2026, "Q1", "Mesa roble", "Online", 35000],
        [2026, "Q2", "Silla cedro", "Tienda", 52000],
        [2026, "Q2", "Mesa roble", "Tienda", 37000],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 5)
    add_border_range(ws, 1, len(data) + 1, 1, 5)
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=5).number_format = '#,##0'
    auto_width(ws)


def avanzado_cap9(wb):
    ws = wb.active
    ws.title = "Auditoría"
    ws["A1"] = "Proceso"; ws["B1"] = "Valor Ingresado"; ws["C1"] = "Fórmula"
    ws["D1"] = "Resultado"; ws["E1"] = "Verificado"
    data = [
        ["Ventas totales", 284500, "=SUM(Datos!E2:E100)", None, "PENDIENTE"],
        ["Costo MP", 95000, "=SUMAPRODUCTO(...)", None, "PENDIENTE"],
        ["Mano obra", 52000, "=SUM(MO!B2:B20)", None, "PENDIENTE"],
        ["Gastos fijos", 35000, "=SUM(Gastos!C2:C15)", None, "PENDIENTE"],
        ["Ganancia", 102500, "=B2-B3-B4-B5", None, "PENDIENTE"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 5)
    add_border_range(ws, 1, len(data) + 1, 1, 5)
    auto_width(ws)


def avanzado_cap10(wb):
    ws = wb.active
    ws.title = "Proyecto Final"
    ws["A1"] = "Componente"; ws["B1"] = "Descripción"; ws["C1"] = "Estado"
    ws["D1"] = "Prioridad"
    data = [
        ["Modelo Datos", "Power Pivot con relaciones tablas transaccionales", "Completado", "Alta"],
        ["Medidas DAX", "CALCULATE, SUMX, FILTER para KPI", "Completado", "Alta"],
        ["Power Query", "Limpieza y transformación de datos fuente", "Completado", "Alta"],
        ["Dashboard", "Controles de formulario + gráficos dinámicos", "En revisión", "Media"],
        ["Power BI", "Publicación en nube del informe completo", "Pendiente", "Media"],
        ["Auditoría", "Validación de fórmulas y modelo", "En revisión", "Alta"],
    ]
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    style_header(ws, 1, 4)
    add_border_range(ws, 1, len(data) + 1, 1, 4)
    auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════
#  GENERACIÓN
# ═══════════════════════════════════════════════════════════════════════

CONFIGS = {
    "excel_basico": [
        ("cap1", "Interfaz y Navegación", basico_cap1),
        ("cap2", "Tipos de Datos", basico_cap2),
        ("cap3", "Fórmulas Básicas", basico_cap3),
        ("cap4", "Referencias", basico_cap4),
        ("cap5", "Formato Celdas", basico_cap5),
        ("cap6", "Funciones", basico_cap6),
        ("cap7", "Gráficos", basico_cap7),
        ("cap8", "Ordenar y Filtrar", basico_cap8),
        ("cap9", "Impresión", basico_cap9),
        ("cap10", "Proyecto Final", basico_cap10),
    ],
    "excel_intermedio": [
        ("cap1", "Formato Condicional", intermedio_cap1),
        ("cap2", "Validación Datos", intermedio_cap2),
        ("cap3", "Búsquedas", intermedio_cap3),
        ("cap4", "Tablas Dinámicas", intermedio_cap4),
        ("cap5", "Gráficos Avanzados", intermedio_cap5),
        ("cap6", "Limpieza Datos", intermedio_cap6),
        ("cap7", "Rangos Nombrados", intermedio_cap7),
        ("cap8", "Análisis Hipotético", intermedio_cap8),
        ("cap9", "Dashboard", intermedio_cap9),
        ("cap10", "Proyecto Final", intermedio_cap10),
    ],
    "excel_avanzado": [
        ("cap1", "Power Query", avanzado_cap1),
        ("cap2", "Power Pivot", avanzado_cap2),
        ("cap3", "DAX Básico", avanzado_cap3),
        ("cap4", "VBA y Macros", avanzado_cap4),
        ("cap5", "Matrices Dinámicas", avanzado_cap5),
        ("cap6", "Solver", avanzado_cap6),
        ("cap7", "Dashboard Avanzado", avanzado_cap7),
        ("cap8", "Power BI", avanzado_cap8),
        ("cap9", "Auditoría", avanzado_cap9),
        ("cap10", "Proyecto Final", avanzado_cap10),
    ],
}

def main():
    for libro, capitulos in CONFIGS.items():
        carpeta = BASE / libro / "codigos"
        carpeta.mkdir(parents=True, exist_ok=True)
        print(f"\n=== {libro} ===")
        for nombre, titulo, fn in capitulos:
            wb = Workbook()
            fn(wb)
            ruta = carpeta / f"{nombre}.xlsx"
            wb.save(ruta)
            print(f"  OK {nombre}.xlsx - {titulo}")

    print(f"\nOK - Generados en:")
    for libro in CONFIGS:
        print(f"   {BASE / libro / 'codigos/'}")

if __name__ == "__main__":
    main()
